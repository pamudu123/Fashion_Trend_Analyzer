from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import os
import args


def add_record_to_excel(track_id, detected_date,detected_time, predicted_age, predicted_gender, upper_custom, bottom_custom):
    excel_file = args.SAVE_EXCEL_RECORDS
    
    if os.path.isfile(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        # Create a new workbook if the file doesn't exist
        workbook = Workbook()
        sheet = workbook.active

        # Make the column names bold
        header = ["Track ID","Detected Date", "Detected Time", "Predicted Age", "Predicted Gender", "Upper Custom Color","Upper Custom Color Percentage", "Bottom Custom Color","Bottom Custom Color Percentage"]
        sheet.append(header)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    upper_custom_color , upper_custom_percentage = upper_custom
    bottom_custom_color , bottom_custom_percentage = bottom_custom
    
    # Define the record data
    record = [track_id, detected_date, detected_time, predicted_age, predicted_gender, upper_custom_color,upper_custom_percentage, bottom_custom_color,bottom_custom_percentage]

    # Append the record to the sheet
    sheet.append(record)

    workbook.save(excel_file)

    print(f"Record added to {excel_file}")

